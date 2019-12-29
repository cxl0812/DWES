import sys
from antlr4 import *
from ScratchLexer import ScratchLexer
from ScratchParser import ScratchParser
from ScratchListener import ScratchListener
import zipfile
import os
import codecs
import xlwt
# workbook相关
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
# 一个eggache的数字转为列字母的方法
from openpyxl.utils import get_column_letter
from openpyxl.reader.excel import load_workbook
import time
from collections import Counter
class HandleExcel():
    def __init__(self):
        self.head_row_labels = [u'Name', u'Abstraction', u'Parallelism', u'LogicalThinking', u'Synchronization',
                                u'FlowControl', u'UserInteractivity', u'DataRepresentation', u'CodeOrganization']
        # self.head_row_labels = [u'Name', u'Abstraction', u'Parallelism', u'LogicalThinking', u'Synchronization',
        #                         u'FlowControl', u'UserInteractivity', u'DataRepresentation', u'CodeOrganization',
        #                         u'n1', u'n2', u'N1', u'N2', u'CC']
    def read_from_file(self):
        info = {}

        path = os.path.abspath('.')
        filepath = path + '/test'
        pathDir = os.listdir(filepath)
        for allDir in pathDir:
            child = os.path.join(filepath, allDir)
            print("name=", allDir)
            listener_info = ctAnalysis(child)
            print("listener_info", listener_info)
            # listener_operators = gen(child)[3]
            # # print("listener_operators", listener_operators)
            # listener_operand = gen(child)[4]
            # # print("listener_operand", listener_operand)
            # listener_cc = gen(child)[5]
            # # print("listener_CC", listener_cc)
            #
            # n1 = len(set(listener_operators))
            # # print("set listener_operators", set(listener_operators))
            # N1 = len(listener_operators)
            # n2 = len(set(listener_operand))
            # # print("set listener_operand", set(listener_operand))
            # N2 = str(len(listener_operand))
            #
            # listener_info['n1'] = n1
            # listener_info['n2'] = n2
            # listener_info['N1'] = N1
            # listener_info['N2'] = N2
            #
            # listener_info['cc'] = listener_cc

            info[allDir] = listener_info
            # print("info[allDir]", info[allDir])

        return info

    def write_to_excel_with_openpyxl(self, records, head_row, save_excel_name):
        # 新建一个workbook
        wb = Workbook()
        # 新建一个excelWriter
        # ew = ExcelWriter(workbook=wb)
        # 设置文件输出路径与名称
        dest_filename = save_excel_name
        # 第一个sheet是ws
        ws = wb.worksheets[0]
        # 设置ws的名称
        ws.title = "range names"
        # 写第一行，标题行
        for h_x in range(1, len(head_row) + 1):
            h_col = get_column_letter(h_x)
            ws.cell('%s%s' % (h_col, 1)).value = '%s' % (head_row[h_x - 1])
        # 写第二行及其以后的那些行
        row = 2
        for name in records:
            ws.cell('%s%s' % ('A', row)).value = '%s' % name
            col = 2
            for point in records[name]:
                col_num = get_column_letter(col)
                ws.cell('%s%s' % (col_num, row)).value = '%s' % records[name][point]
                col += 1
            row += 1
        # ew.save(filename=dest_filename)
        wb.save(filename=dest_filename)


    def run_main_save_to_excel_with_openpyxl(self):
        dataset_list = self.read_from_file()
        '''test use openpyxl to handle EXCEL 2007'''
        head_row_label = self.head_row_labels
        save_name = "test_openpyxl.xlsx"
        self.write_to_excel_with_openpyxl(dataset_list, head_row_label, save_name)
def unzip_scratch(filename):
    """
    unzip scratch project and extract project.json file
    :param filename: filename fo scratch project
    :return: null or project.json content
    """
    zfile = zipfile.ZipFile(filename, 'r')
    if "project.json" in zfile.namelist():
        data = zfile.read("project.json")
        return data
    else:
        return None

def ctAnalysis(argv):
    raw_json = unzip_scratch(argv)
    encoded_json = codecs.decode(raw_json, 'utf-8', 'strict')
    input = InputStream(encoded_json)
    lexer = ScratchLexer(input)
    stream = CommonTokenStream(lexer)
    parser = ScratchParser(stream)
    tree = parser.json()
    walker = ParseTreeWalker()
    listener = ScratchListener()
    walker.walk(listener, tree)
    ct_score = {}
    ct_score['abstract'] = 0
    ct_score['parallel'] = 0
    ct_score['logic'] = 0
    ct_score['synchron'] = 0
    ct_score['flowControl'] = 0
    ct_score['userInteraction'] = 0
    ct_score['dataRepresent'] = 1
    ct_score['codeOrganize'] = 0
    block_count = listener.block_count  # block数目
    sprite_count = listener.sprite_count
    comment_count = listener.comment_count  # 注释数目
    block = listener.id_block  # block
    broadcast_received = listener.broadcast_received
    broadcast = listener.broadcast
    listdic = listener.list
    # print('listdic',listdic)
    listlen = {}
    for i in listdic.keys():
        # print('i!!!!', i)
        listlen[i] = len(listdic[i])
    # print('listlen!!!!',listlen)
    isQueue = listdic.copy()
    # print('isQueue!!!!',isQueue)
    isStack = listdic.copy()
    deadBlock_count = 0
    flagClicked_count = 0
    keyPressed_count = 0
    spriteClicked_count = 0
    backdropSwitch_count = 0
    backClicked_count = 0
    greaterThan_count = 0
    broadcastRecive_dicc = Counter()
    motionGreater_count = 0
    logic_operation = {'operator_and', 'operator_or', 'operator_not'}
    sprite_init = 1

    # 遍历所有block，给出ct得分
    for id in block.keys():
        isDead_code = block[id].getIsDead()
        name = block[id].getName()

        parent_id = block[id].getParent()
        # 父block是deadcode则子block一定是deadcode
        if parent_id is not None and parent_id in block.keys() and block[parent_id].getIsDead() is True:
            block[id].setIsDead(True)
            deadBlock_count += 1
            substack = block[id].getSubstack()
            if substack is not None and substack in block.keys():
                block[substack].setIsDead(True)
                deadBlock_count += 1
            substack2 = block[id].getSubstack2()
            if substack2 is not None and substack2 in block.keys():
                block[substack2].setIsDead(True)
                deadBlock_count += 1
            next_id = block[id].getNext()
            while next_id is not None and next_id in block.keys():
                block[next_id].setIsDead(True)
                deadBlock_count += 1
                substack = block[next_id].getSubstack()
                if substack is not None and substack in block.keys():
                    block[substack].setIsDead(True)
                    deadBlock_count += 1
                substack2 = block[id].getSubstack2()
                if substack2 is not None and substack2 in block.keys():
                    block[substack2].setIsDead(True)
                    deadBlock_count += 1
                next_id = block[next_id].getNext()
        # 子block继承父block的proc
        if parent_id is not None and parent_id in block.keys() and block[parent_id].getProcId() is not None and block[id].getProcId() is not None:
            proc_id = block[parent_id].getProcId()
            block[id].setProcId(proc_id)
            substack = block[id].getSubstack()
            if substack is not None and substack in block.keys():
                block[substack].setProcId(proc_id)
            substack2 = block[id].getSubstack2()
            if substack2 is not None and substack2 in block.keys():
                block[substack2].setProcId(proc_id)
            next_id = block[id].getNext()
            while next_id is not None and next_id in block.keys():
                block[next_id].setProcId(proc_id)
                substack = block[next_id].getSubstack()
                if substack is not None and substack in block.keys():
                    block[substack].setProcId(proc_id)
                substack2 = block[id].getSubstack2()
                if substack2 is not None and substack2 in block.keys():
                    block[substack2].setProcId(proc_id)
                next_id = block[next_id].getNext()
        # 判断是否为deadcode
        if isDead_code is False:
            if parent_id is None:  # 判断block的起始语句是否为启动语句
                if name.find("when") == -1 and name.find("start") == -1 and name.find("definition") == -1:
                    block[id].setIsDead(True)
                    deadBlock_count += 1
                    substack = block[id].getSubstack()
                    if substack is not None and substack in block.keys():
                        block[substack].setIsDead(True)
                        deadBlock_count += 1
                    substack2 = block[id].getSubstack2()
                    if substack2 is not None and substack2 in block.keys():
                        block[substack2].setIsDead(True)
                        deadBlock_count += 1
                    # 将所有的next block设置为deadcode
                    next_id = block[id].getNext()
                    while next_id is not None and next_id in block.keys():
                        block[next_id].setIsDead(True)
                        deadBlock_count += 1
                        substack = block[next_id].getSubstack()
                        if substack is not None and substack in block.keys():
                            block[substack].setIsDead(True)
                            deadBlock_count += 1
                        substack2 = block[next_id].getSubstack2()
                        if substack2 is not None and substack2 in block.keys():
                            block[substack2].setIsDead(True)
                            deadBlock_count += 1
                        next_id = block[next_id].getNext()
        # proc记录
        if name == 'procedures_definition':
            block[id].setProcId(id)
            substack = block[id].getSubstack()
            if substack is not None and substack in block.keys():
                block[substack].setProcId(id)
            substack2 = block[id].getSubstack2()
            if substack2 is not None and substack2 in block.keys():
                block[substack2].setProcId(id)
            next_id = block[id].getNext()
            while next_id is not None and next_id in block.keys():
                block[next_id].setProcId(id)
                substack = block[next_id].getSubstack()
                if substack is not None and substack in block.keys():
                    block[substack].setProcId(id)
                substack2 = block[id].getSubstack2()
                if substack2 is not None and substack2 in block.keys():
                    block[substack2].setProcId(id)
                next_id = block[next_id].getNext()
        if isDead_code is False:
            if name == 'event_whengreaterthan':
                greaterThan_count += 1
            if name == 'videoSensing_whenMotionGreaterThan':
                motionGreater_count += 1
            if name == 'event_whenbroadcastreceived' and block[id].getBroadcast() in broadcast:
                # 评分标准 4-5
                if ct_score['synchron'] < 5:
                    ct_score['synchron'] = 5
                broadcastRecive_dicc[block[id].getBroadcast()] += 1

            if name == 'event_whenkeypressed':
                # 评分标准4-2
                if ct_score['synchron'] < 2:
                    ct_score['synchron'] = 2
                # 评分标准6-3
                if ct_score['userInteraction'] < 3:
                    ct_score['userInteraction'] = 3
                keyPressed_count += 1
            if name == 'event_whenthisspriteclicked':
                # 评分标准4-2
                if ct_score['synchron'] < 2:
                    ct_score['synchron'] = 2
                # 评分标准6-3
                if ct_score['userInteraction'] < 3:
                    ct_score['userInteraction'] = 3
                spriteClicked_count += 1
            if name == 'event_whenbackdropswitchesto':
                backdropSwitch_count += 1
                # 评分标准4-2
                if ct_score['synchron'] < 2:
                    ct_score['synchron'] = 2
            if name == 'event_whenstageclicked':
                # 评分标准4-2
                if ct_score['synchron'] < 2:
                    ct_score['synchron'] = 2
                # 评分标准6-3
                if ct_score['userInteraction'] < 3:
                    ct_score['userInteraction'] = 3
                backClicked_count += 1
            if name == 'event_whenflagclicked':
                # 评分标准6-2
                if ct_score['userInteraction'] < 2:
                    ct_score['userInteraction'] = 2
                flagClicked_count += 1

            # 列表操作
            if name == 'data_deleteoflist':
                loc = block[id].getLocation()
                if type(loc) is int:
                    listname = block[id].getListName()
                    if (listname in isQueue.keys()) and loc != 1:
                        del isQueue[listname]

                    if (listname in isStack.keys()) and loc != listlen[listname]:
                        del isStack[listname]

                    if listname in listdic.keys() and listlen[listname] >= loc:
                        listdic[listname].pop(loc - 1)
                        # print("删除list", listname, loc, listdic[listname ])
                        listlen[listname] -= 1
            if name == 'data_addtolist':
                # print('list', listdic.items())
                listname = block[id].getListName()
                # print('listname', listname)
                content = block[id].getListContent()
                if listname in listdic.keys():
                    listdic[listname].append(content)
                    listlen[listname] += 1
                    # print("添加list", listname, listdic[listname])
            if name == 'data_insertatlist':
                loc = block[id].getLocation()
                if type(loc) is int:
                    listname = block[id].getListName()
                    content = block[id].getListContent()
                    if (listname in isQueue.keys()) and loc != listlen[listname] + 1:
                        del isQueue[listname]

                    if (listname in isStack.keys()) and loc != listlen[listname] + 1:
                        del isStack[listname]

                    if listname in listdic.keys() and loc <= listlen[listname]:
                        listdic[listname].insert(loc - 1, content)
                        # print("插入list", listname, loc, listdic[listname])
                        listlen[listname] += 1
            if name == 'data_replaceitemoflist':
                loc = block[id].getLocation()

                if type(loc) is int:
                    listname = block[id].getListName()
                    # print("替换list！！！", listname, loc, listlen[listname], listdic[listname])
                    content = block[id].getListContent()
                    if listname in isQueue.keys():
                        del isQueue[listname]

                    if listname in isStack.keys():
                        del isStack[listname]

                    if listname in listdic.keys() and loc <= listlen[listname]:
                        listdic[listname][loc - 1] = content
                        # print("替换list", listname, loc, listdic[listname])

            # 角色属性初始化
            if name == 'looks_nextcostume':
                isfind = 0
                i = 0
                while isfind == 0 and parent_id is not None and parent_id in block.keys():

                    if block[parent_id].getName() == 'looks_switchcostumeto':
                        isfind = 1
                    parent_id = block[parent_id].getParent()
                    i += 1
                if isfind == 0:
                    sprite_init = 0
            if name == 'looks_changesizeby':
                isfind = 0
                while isfind == 0 and parent_id is not None and parent_id in block.keys():
                    if block[parent_id].getName() == 'looks_setsizeto':
                        isfind = 1
                    parent_id = block[parent_id].getParent()
                if isfind == 0:
                    sprite_init = 0
            if name == 'looks_nextbackdrop':
                isfind = 0
                while isfind == 0 and parent_id is not None and parent_id in block.keys():
                    if block[parent_id].getName() == 'looks_switchbackdropto':
                        isfind = 1
                    parent_id = block[parent_id].getParent()
                if isfind == 0:
                    sprite_init = 0
            if name in {'motion_turnright', 'motion_turnleft', 'motion_pointtowards'}:
                isfind = 0
                while isfind == 0 and parent_id is not None and parent_id in block.keys():
                    if block[parent_id].getName() == 'motion_pointindirection':
                        isfind = 1
                    parent_id = block[parent_id].getParent()
                if isfind == 0:
                    sprite_init = 0
            if name in {'motion_movesteps', 'motion_glideto', 'motion_glidesecstoxy', 'motion_goto'}:
                isfind = 0
                while isfind == 0 and parent_id is not None and parent_id in block.keys():
                    if block[parent_id].getName() == 'motion_gotoxy':
                        isfind = 1
                    parent_id = block[parent_id].getParent()
                if isfind == 0:
                    sprite_init = 0

            # 评分标准1-5
            if ct_score['abstract'] < 5 and name == 'procedures_call' and block[id].getProcId() is not None:
                ct_score['abstract'] = 5
            # 评分标准1-4
            if ct_score['abstract'] < 4 and name.find('clone') != -1:
                ct_score['abstract'] = 4
            # 评分标准1-3
            if ct_score['abstract'] < 3 and name == 'procedures_call':
                ct_score['abstract'] = 3
            # 评分标准1-2
            if ct_score['abstract'] < 2 and name in{'looks_switchbackdroptoandwait', 'looks_backdrops',
                                                    'looks_nextbackdrop', 'looks_nextcostume', 'looks_switchcostumeto'} :
                ct_score['abstract'] = 2

            if name == 'control_if' or name == 'control_if_else':
                substack = block[id].getSubstack()
                isfind = 0
                while isfind == 0 and substack is not None and substack in block.keys():
                    if block[substack].getName() in {'control_if', 'control_if_else'}:
                        isfind = 1
                    if block[substack].getName() in {'control_repeat', 'control_forever', 'control_repeat_until'}:
                        isfind = 2
                    substack = block[substack].getNext()
                # 评分标准3-5
                if ct_score['logic'] < 5 and isfind == 2:
                    ct_score['logic'] = 5
                # 评分标准3-4
                if ct_score['logic'] < 4 and isfind == 1:
                    ct_score['logic'] = 4
                # 评分标准3-3
                if ct_score['logic'] < 3:
                    condition = block[id].getCondition()
                    if condition in logic_operation:
                        ct_score['logic'] = 3
            if name == 'control_if_else':
                substack2 = block[id].getSubstack2()
                isfind = 0
                while isfind == 0 and substack2 is not None and substack2 in block.keys():
                    if block[substack2].getName() in {'control_if', 'control_if_else'}:
                        isfind = 1
                    if block[substack2].getName() in {'control_repeat', 'control_forever', 'control_repeat_until'}:
                        isfind = 2
                    substack2 = block[substack2].getNext()
                # 评分标准3-5
                if ct_score['logic'] < 5 and isfind == 2:
                    ct_score['logic'] = 5
                # 评分标准3-4
                if ct_score['logic'] < 4 and isfind == 1:
                    ct_score['logic'] = 4

            # 评分标准3-1
            if ct_score['logic'] == 0 and name == 'control_if':
                ct_score['logic'] = 1
            # 评分标准3-2
            if ct_score['logic'] < 2 and name == 'control_if_else':
                ct_score['logic'] = 2

            # 评分标准4-4
            if ct_score['synchron'] < 4 and name == 'control_wait_until':
                ct_score['synchron'] = 4
            # 评分标准4-3
            if ct_score['synchron'] < 3 and name in {'sensing_touchingcolor', 'sensing_coloristouchingcolor',
                                                     'sensing_loudness', 'sensing_timer', 'sensing_current',
                                                     'videoSensing_videoOn', 'sensing_dayssince2000', 'sensing_of'}:
                ct_score['synchron'] = 3
            # 评分标准4-2
            if ct_score['synchron'] < 2 and name in {'sensing_touchingobject', 'sensing_keypressed', 'sensing_mousedown',
                                                     'sensing_mousex', 'sensing_mousey', 'sensing_distanceto', 'looks_costumenumbername'}:
                ct_score['synchron'] = 2
            # 评分标准4-1
            if ct_score['synchron'] == 0 and name in {'control_wait', 'control_stop'}:
                ct_score['synchron'] = 1

            if name in {'control_repeat', 'control_forever', 'control_repeat_until'}:
                substack = block[id].getSubstack()
                isfind = 0
                while isfind == 0 and substack is not None and substack in block.keys():
                    if block[substack].getName() in {'control_if', 'control_if_else', 'control_repeat',
                                                     'control_forever', 'control_repeat_until'}:
                        isfind = 1
                    substack = block[substack].getNext()
                # 评分标准5-5
                if ct_score['flowControl'] < 5 and isfind == 1:
                    ct_score['flowControl'] = 5
                # 评分标准5-4
                if ct_score['flowControl'] < 4:
                    condition = block[id].getCondition()
                    if condition in logic_operation:
                        ct_score['flowControl'] = 4
            if name == 'control_repeat_until':
                # 评分标准5-3
                if ct_score['flowControl'] < 3:
                    ct_score['flowControl'] = 3
            # 评分标准5-2
            if ct_score['flowControl'] < 2 and name in {'control_repeat', 'control_forever'}:
                ct_score['flowControl'] = 2

            # 评分标准6-5
            if ct_score['userInteraction'] < 5 and name == 'sensing_askandwait':
                ct_score['userInteraction'] = 5
                # 评分标准6-4
            if ct_score['userInteraction'] < 4 and name in {'videoSensing_whenMotionGreaterThan',
                                                            'videoSensing_videoOn', 'sensing_loudness'}:
                ct_score['userInteraction'] = 4
            # 评分标准6-3
            if ct_score['userInteraction'] < 3 and name in {'sensing_keypressed', 'sensing_mousedown'}:
                ct_score['userInteraction'] = 3
            # 评分标准6-1
            if ct_score['userInteraction'] == 0 and name in {'looks_sayforsecs', 'looks_say', 'looks_thinkforsecs', 'looks_think'}:
                ct_score['userInteraction'] = 1

            # 评分标准7-4
            if ct_score['dataRepresent'] < 4 and name in {'data_addtolist', 'data_deleteoflist',
                                                          'data_deletealloflist', 'data_insertatlist',
                                                          'data_replaceitemoflist', 'data_itemoflist',
                                                          'data_itemnumoflist', 'data_lengthoflist',
                                                          'data_listcontainsitem', 'data_showlist',
                                                          'data_hidelist'}:
                ct_score['dataRepresent'] = 4
            # 评分标准7-3
            if ct_score['dataRepresent'] < 3 and name in {'data_setvariableto', 'data_changevariableby',
                                                          'data_showvariable', 'data_hidevariable'}:
                ct_score['dataRepresent'] = 3
            # 评分标准7-2
            if ct_score['dataRepresent'] < 2 and name == 'operator_join':
                ct_score['dataRepresent'] = 2

    # 评分标准1-1
    if ct_score['abstract'] == 0 and sprite_count > 1 and block_count > 1:
        ct_score['abstract'] = 1
    # 评分标准5-1
    if ct_score['flowControl'] == 0 and block_count > 1:
        ct_score['flowControl'] = 1
    # 评分标准2-5
    for key in broadcastRecive_dicc.keys():
        if broadcastRecive_dicc[key] > 1 and ct_score['parallel'] < 5:
            ct_score['parallel'] = 5
            break
    # 评分标准2-4
    if ct_score['parallel'] < 4 and (greaterThan_count > 1 or motionGreater_count > 1):
        ct_score['parallel'] = 4
    # 评分标准2-3
    if ct_score['parallel'] < 3 and backdropSwitch_count > 1:
        ct_score['parallel'] = 3
    # 评分标准2-2
    if ct_score['parallel'] < 2 and (keyPressed_count > 1 or spriteClicked_count > 1 or backClicked_count > 1):
        ct_score['parallel'] = 2
    # 评分标准2-1
    if ct_score['parallel'] == 0 and flagClicked_count > 1:
        ct_score['parallel'] = 1

    # 评分标准7-5
    if ct_score['dataRepresent'] < 5 and (len(isQueue) != 0 or len(isStack) != 0):
        ct_score['dataRepresent'] = 5

    if broadcast_received and broadcast:
        isfind = 1
        for message in broadcast:
            if message not in broadcast_received:
                isfind = 0
                break
        # 评分标准8-5
        if ct_score['codeOrganize'] < 5 and isfind == 1:
            ct_score['codeOrganize'] = 5
    # 评分标准8-4
    if ct_score['codeOrganize'] < 4 and deadBlock_count == 0:
        ct_score['codeOrganize'] = 4
    # 评分标准8-3
    if ct_score['codeOrganize'] < 3 and comment_count > 0:
        ct_score['codeOrganize'] = 3
    # 评分标准8-1
    if ct_score['codeOrganize'] == 0 and sprite_init == 1:
        ct_score['codeOrganize'] = 1

    return ct_score


if __name__ == '__main__':
    start_time = time.time()
    p = HandleExcel()
    p.run_main_save_to_excel_with_openpyxl()
    # ctAnalysis(sys.argv[1])
    end_time = time.time()
    spend_time = end_time - start_time
    print("耗时：%.4f" % spend_time)
